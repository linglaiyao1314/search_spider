# coding=utf-8
from engin.main import bdcrawl, momocrawl, pinglecrawl, yitaocrawl
import json
from flask import Flask, request
from engin.logs import search_logger, INFO, DEBUG, ERROR, wrapstring
from openpyxl import Workbook, load_workbook
import re
import random
import string

app = Flask(__name__)
app.debug = False


# HEADERS = {"User-Agent":
#                "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko)Chrome/38.0.2125.122 Safari/537.36"}
HEADERS = {"User-Agent":
               "Mozilla/5.0 (Windows NT 6.1; WOW64; rv:39.0) Gecko/20100101 Firefox/39.0"}
STATE = 0

from functools import wraps
from flask import make_response


def allow_cross_domain(fun):
    @wraps(fun)
    def wrapper_fun(*args, **kwargs):
        rst = make_response(fun(*args, **kwargs))
        rst.headers['Access-Control-Allow-Origin'] = '*'
        rst.headers['Access-Control-Allow-Methods'] = 'PUT,GET,POST,DELETE'
        allow_headers = "Referer,Accept,Origin,User-Agent"
        rst.headers['Access-Control-Allow-Headers'] = allow_headers
        return rst
    return wrapper_fun


class MyWorkBook(object):
    def __init__(self, bookname=False):
        if bookname:
            try:
                self.book = load_workbook(bookname)
            except:
                self.book = Workbook()
                self.book.save(bookname)
        else:
            self.book = Workbook()
        self.default_sheet = self.book.active

    def create_sheet(self, index=None, title=None):
        return self.book.create_sheet(index, title)

    def get_sheet_by_name(self, name):
        try:
            return self.book[name]
        except:
            return self.create_sheet(title=name)

    def save(self, name):
        self.book.save(name)

    def insert_row(self, datas, sheet=None):
        if sheet:
            isheet = self.get_sheet_by_name(sheet)
        else:
            isheet = self.default_sheet
        if not isinstance(datas, list):
            isheet.append([datas])
        else:
            isheet.append(datas)

def activity_api(itemlist):
    for item in itemlist:
        if u"图书" in item[-1]:
            item[-1] = u"图书"
    return itemlist


@app.route('/', methods=['POST', 'GET'])
@app.route('/index/', methods=["POST", 'GET'])
def json_result():
    search_logger.info(".........START to ..........Search...................\n")
    keywords, country_id = ("", 1)
    if request.method == 'POST':
        keywords = request.form.get("keywords", "")
        country_id = int(request.form.get("country_id", 1))
    elif request.method == 'GET':
        keywords = request.args.get("keywords", None)
        country_id = int(request.args.get("country_id", 1))
    search_logger.info(wrapstring("Keywords is : [ %s ]" % keywords))
    if country_id == 1:
        result = bdcrawl(keywords, headers=HEADERS, limit=10)
    elif country_id == 2:
        result = pinglecrawl(keywords, headers=HEADERS, limit=10)
        if not result:
            search_logger.info(wrapstring("Pingle have no result, so go to momo and pchome...."))
            result = momocrawl(keywords, headers=HEADERS, limit=10)
            # result = momo_pc_event(keywords, headers=HEADERS, shop=shop)
    else:
        result = []
    search_logger.debug(wrapstring("There %d items will be return\n\n" % len(result), DEBUG))
    api_result = {"data": []}
    for every_result in result:
        dct = {}
        # "good_name", "price", "image_url", "good_url"
        dct["product_name"] = every_result[1]
        dct["url"] = every_result[4]
        shop = every_result[0]
        if shop == 1:
            dct["shop"] = u"京东"
        elif shop == 8:
            dct["shop"] = u"苏宁"
        elif shop == 3:
            dct["shop"] = u"一号店"
        elif shop == 4:
            dct["shop"] = u"天猫"
        elif shop == 5:
            dct["shop"] = u"当当"
        dct["price"] = every_result[2]
        dct["product_img"] = every_result[3]
        api_result["data"].append(dct)
    return json.dumps(api_result)


def get_number(s):
    pattern = re.compile(u"[\d]")
    return filter(lambda x: re.match(pattern, x), s)


@app.route('/excel', methods=["POST", 'GET'])
@allow_cross_domain
def write_excel():
    global STATE
    if request.method == 'POST':
        row_dct = json.loads(request.form.get("data"))
        bookname = r"C:\Users\admin\Desktop\excelXXXXX\source.xlsx"
        w = MyWorkBook(bookname)
        if STATE == 0:
            titles = [u"商品名", u"价格", u"原始价格", u"商家编号", u"产品网址", "opt_id"]
            w.insert_row(titles)
            STATE += 1
        row = []
        code = get_number(row_dct["url"])[:10]
        row.append(row_dct["product_name"])
        row.append(row_dct["price"])
        row.append(row_dct["price"])
        row.append(code)
        row.append(row_dct["url"])
        row.append(45)
        w.insert_row(row)
        w.save(bookname)
        print row
        return "ok"
    elif request.method == "GET":
        print "hello"
        return "ok"

if __name__ == '__main__':
    app.run("0.0.0.0", 8888, debug=False)
    # app.run()
